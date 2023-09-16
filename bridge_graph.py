import math
import matplotlib.pyplot as plt
import numpy as np
import sympy as sp
from function import stop_watch
import datetime
import openpyxl as xl
import time

# TODO 信頼度　m3、m4を出力する
# TODO 完全グラフ

@stop_watch
def f_actual(x):
    λa = 0.01
    λb = 0.01
    λc = 0.01
    λd = 0.01
    λe = 0.01
    a = sp.Symbol('a')
    b = sp.Symbol('b')
    c = sp.Symbol('c')
    d = sp.Symbol('d')
    e = sp.Symbol('e')
    exp = sp.E
    Rt = (1 - a) * b * c * (1 - d) * (1 - e) + a * b * c * (1 - d) * (1 - e) + a * (1 - b) * (1 - c) * d * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (1 - a) * b * c * d * (1 - e) + a * b * c * d * (1 - e) + a * (1 - b) * c * (1 - d) * e + (1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (1 - a) * b * c * d * e + a * b * c * d * e
    actual_Rt = Rt.subs([(a, exp ** (-λa * x)), (b, exp ** (-λb * x)), (c, exp ** (-λc * x)), (d, exp ** (-λd * x)),(e, exp ** (-λe * x))])
    a = time.time()
    diff_Rt = sp.diff(actual_Rt, x)
    print(f"微分時間：{time.time()-a}")
    actual = (-1 / actual_Rt) * diff_Rt
    return actual

@stop_watch
def f_3(t):
    # 初期値入力
    λa = 0.01
    λb = 0.01
    λc = 0.01
    λd = 0.01
    λe = 0.01
    exp = math.e
    a = exp ** (-λa * t)
    b = exp ** (-λb * t)
    c = exp ** (-λc * t)
    d = exp ** (-λd * t)
    e = exp ** (-λe * t)
    R_diff = [-λa * exp ** (-λa * t),
              -λb * exp ** (-λb * t),
              -λc * exp ** (-λc * t),
              -λd * exp ** (-λd * t),
              -λe * exp ** (-λe * t)]
    link_list = ['a', 'b', 'c', 'd', 'e']
    round_dic = {}
    result_dic = {}
    num = [1, 0]
    Rt_min_3 = a*b*c*(1-d)*(1-e)+a*b*(1-c)*d*(1-e)+a*(1-b)*c*d*(1-e)+(1-a)*b*c*d*(1-e)+\
            a*b*c*d*(1-e)+a*(1-b)*c*(1-d)*e+(1-a)*b*c*(1-d)*e+a*b*c*(1-d)*e+a*(1-b)*(1-c)*d*e+\
            (1-a)*b*(1-c)*d*e+a*b*(1-c)*d*e+a*(1-b)*c*d*e+(1-a)*b*c*d*e+a*b*c*d*e
    Rt_Max_3 = 1 - (a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e)

    """
    偏微分計算。
    a~eそれぞれに1、0を代入した値を取得（下限と上限ごとに）
    """
    for a in num:
        min_3 = a*b*c*(1-d)*(1-e)+a*b*(1-c)*d*(1-e)+a*(1-b)*c*d*(1-e)+\
                (1-a)*b*c*d*(1-e)+a*b*c*d*(1-e)+a*(1-b)*c*(1-d)*e+(1-a)*b*c*(1-d)*e+\
                a*b*c*(1-d)*e+a*(1-b)*(1-c)*d*e+(1-a)*b*(1-c)*d*e+a*b*(1-c)*d*e+\
                a*(1-b)*c*d*e+(1-a)*b*c*d*e+a*b*c*d*e
        max_3 = 1 - (a*b*(1-c)*(1-d)*e+(1-a)*(1-b)*c*d*e)
        round_dic[f'+{a}'] = max_3
        round_dic[f'-{a}'] = min_3
        if a == 0:
            # 偏微分の部分の上限(fMax(1)-fmin(0))を導出
            max_round = round_dic['+1'] - round_dic['-0']
            # 偏微分の部分の下限(fmin(1)-fMax(0))を導出　　この値が常に正とは限らない。
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['a'] = [max_round, min_round]
    for b in num:
        a = exp ** (-λa * t)
        min_3 = a*b*c*(1-d)*(1-e)+a*b*(1-c)*d*(1-e)+a*(1-b)*c*d*(1-e)+\
                (1-a)*b*c*d*(1-e)+a*b*c*d*(1-e)+a*(1-b)*c*(1-d)*e+(1-a)*b*c*(1-d)*e+\
                a*b*c*(1-d)*e+a*(1-b)*(1-c)*d*e+(1-a)*b*(1-c)*d*e+a*b*(1-c)*d*e+\
                a*(1-b)*c*d*e+(1-a)*b*c*d*e+a*b*c*d*e
        max_3 = 1 - (a*b*(1-c)*(1-d)*e+(1-a)*(1-b)*c*d*e)
        round_dic[f'+{b}'] = max_3
        round_dic[f'-{b}'] = min_3
        if b == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['b'] = [max_round, min_round]
    for c in num:
        b = exp ** (-λb * t)
        min_3 = a*b*c*(1-d)*(1-e)+a*b*(1-c)*d*(1-e)+a*(1-b)*c*d*(1-e)+\
                (1-a)*b*c*d*(1-e)+a*b*c*d*(1-e)+a*(1-b)*c*(1-d)*e+(1-a)*b*c*(1-d)*e+\
                a*b*c*(1-d)*e+a*(1-b)*(1-c)*d*e+(1-a)*b*(1-c)*d*e+a*b*(1-c)*d*e+\
                a*(1-b)*c*d*e+(1-a)*b*c*d*e+a*b*c*d*e
        max_3 = 1 - (a*b*(1-c)*(1-d)*e+(1-a)*(1-b)*c*d*e)
        round_dic[f'+{c}'] = max_3
        round_dic[f'-{c}'] = min_3
        if c == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['c'] = [max_round, min_round]
    for d in num:
        c = exp ** (-λc * t)
        min_3 = a*b*c*(1-d)*(1-e)+a*b*(1-c)*d*(1-e)+a*(1-b)*c*d*(1-e)+\
                (1-a)*b*c*d*(1-e)+a*b*c*d*(1-e)+a*(1-b)*c*(1-d)*e+(1-a)*b*c*(1-d)*e+\
                a*b*c*(1-d)*e+a*(1-b)*(1-c)*d*e+(1-a)*b*(1-c)*d*e+a*b*(1-c)*d*e+\
                a*(1-b)*c*d*e+(1-a)*b*c*d*e+a*b*c*d*e
        max_3 = 1 - (a*b*(1-c)*(1-d)*e+(1-a)*(1-b)*c*d*e)
        round_dic[f'+{d}'] = max_3
        round_dic[f'-{d}'] = min_3
        if d == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['d'] = [max_round, min_round]
    for e in num:
        d = exp ** (-λd * t)
        min_3 = a*b*c*(1-d)*(1-e)+a*b*(1-c)*d*(1-e)+a*(1-b)*c*d*(1-e)+\
                (1-a)*b*c*d*(1-e)+a*b*c*d*(1-e)+a*(1-b)*c*(1-d)*e+(1-a)*b*c*(1-d)*e+\
                a*b*c*(1-d)*e+a*(1-b)*(1-c)*d*e+(1-a)*b*(1-c)*d*e+a*b*(1-c)*d*e+\
                a*(1-b)*c*d*e+(1-a)*b*c*d*e+a*b*c*d*e
        max_3 = 1 - (a*b*(1-c)*(1-d)*e+(1-a)*(1-b)*c*d*e)
        round_dic[f'+{e}'] = max_3
        round_dic[f'-{e}'] = min_3
        if e == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['e'] = [max_round, min_round]
    """
    ↑の操作で、偏微分の値を取得。
    その値と、dR(t)/dtの値を
    掛け合わせたものの和を取る。
    故障率の上限と下限を算出
    """
    diff_Rt_Max = 0
    diff_Rt_min = 0
    for idx, link in enumerate(link_list):
        diff_Rt_Max = diff_Rt_Max + result_dic[link][0] * R_diff[idx]
        diff_Rt_min = diff_Rt_min + result_dic[link][1] * R_diff[idx]
    #(Max/minで故障率上限導出)
    approximation_Max = (-1/Rt_min_3)*diff_Rt_Max
    #(min/Maxで故障率上限導出)
    approximation_min = (-1/Rt_Max_3)*diff_Rt_min
    return approximation_Max, approximation_min, Rt_Max_3, Rt_min_3
@stop_watch
def f_4(t):
    # 初期値入力
    λa = 0.01
    λb = 0.01
    λc = 0.01
    λd = 0.01
    λe = 0.01
    exp = math.e
    a = exp ** (-λa * t)
    b = exp ** (-λb * t)
    c = exp ** (-λc * t)
    d = exp ** (-λd * t)
    e = exp ** (-λe * t)
    R_diff = [-λa * exp ** (-λa * t),
              -λb * exp ** (-λb * t),
              -λc * exp ** (-λc * t),
              -λd * exp ** (-λd * t),
              -λe * exp ** (-λe * t)]
    link_list = ['a', 'b', 'c', 'd', 'e']
    round_dic = {}
    result_dic = {}
    num = [1, 0]
    first_min_4 = a * b * c * (1 - d) * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (
                1 - a) * b * c * d * (1 - e) \
                  + a * b * c * d * (1 - e) + a * b * (1 - c) * (1 - d) * e + a * (1 - b) * c * (1 - d) * e + (
                              1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e \
                  + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (
                              1 - a) * b * c * d * e + a * b * c * d * e
    first_max_4 = 1 -( a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e + a * b * (1 - c) * (1 - d) * (
                1 - e) + a * (1 - b) * c * (1 - d) * (1 - e) + \
                  (1 - a) * b * (1 - c) * d * (1 - e) + (1 - a) * (1 - b) * c * d * (1 - e) + a * (1 - b) * (1 - c) * (
                              1 - d) * e + (1 - a) * b * (1 - c) * (1 - d) * e + \
                  (1 - a) * (1 - b) * c * (1 - d) * e + (1 - a) * (1 - b) * (1 - c) * d * e)

    """
    偏微分計算。
    a~eそれぞれに1、0を代入した値を取得（下限と上限ごとに）
    """
    for a in num:
        min_4 = a * b * c * (1 - d) * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (
                1 - a) * b * c * d * (1 - e) \
                  + a * b * c * d * (1 - e) + a * b * (1 - c) * (1 - d) * e + a * (1 - b) * c * (1 - d) * e + (
                              1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e \
                  + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (
                              1 - a) * b * c * d * e + a * b * c * d * e
        max_4 = 1 - (a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e + a * b * (1 - c) * (1 - d) * (
                    1 - e) + a * (1 - b) * c * (1 - d) * (1 - e) + \
                      (1 - a) * b * (1 - c) * d * (1 - e) + (1 - a) * (1 - b) * c * d * (1 - e) + a * (1 - b) * (1 - c) * (
                                  1 - d) * e + (1 - a) * b * (1 - c) * (1 - d) * e + \
                      (1 - a) * (1 - b) * c * (1 - d) * e + (1 - a) * (1 - b) * (1 - c) * d * e)
        round_dic[f'+{a}'] = max_4
        round_dic[f'-{a}'] = min_4
        if a == 0:
            # 偏微分の部分の上限(fMax(1)-fmin(1))を導出
            max_round = round_dic['+1'] - round_dic['-0']
            # 偏微分の部分の下限(fmin(1)-fMax(1))を導出　　この値が常に正とは限らない。
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['a'] = [max_round, min_round]
    for b in num:
        a = exp ** (-λa * t)
        min_4 = a * b * c * (1 - d) * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (
                1 - a) * b * c * d * (1 - e) \
                  + a * b * c * d * (1 - e) + a * b * (1 - c) * (1 - d) * e + a * (1 - b) * c * (1 - d) * e + (
                              1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e \
                  + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (
                              1 - a) * b * c * d * e + a * b * c * d * e
        max_4 = 1 - (a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e + a * b * (1 - c) * (1 - d) * (
                    1 - e) + a * (1 - b) * c * (1 - d) * (1 - e) + \
                      (1 - a) * b * (1 - c) * d * (1 - e) + (1 - a) * (1 - b) * c * d * (1 - e) + a * (1 - b) * (1 - c) * (
                                  1 - d) * e + (1 - a) * b * (1 - c) * (1 - d) * e + \
                      (1 - a) * (1 - b) * c * (1 - d) * e + (1 - a) * (1 - b) * (1 - c) * d * e)
        round_dic[f'+{b}'] = max_4
        round_dic[f'-{b}'] = min_4
        if b == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['b'] = [max_round, min_round]
    for c in num:
        b = exp ** (-λb * t)
        min_4 = a * b * c * (1 - d) * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (
                1 - a) * b * c * d * (1 - e) \
                  + a * b * c * d * (1 - e) + a * b * (1 - c) * (1 - d) * e + a * (1 - b) * c * (1 - d) * e + (
                              1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e \
                  + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (
                              1 - a) * b * c * d * e + a * b * c * d * e
        max_4 = 1 -( a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e + a * b * (1 - c) * (1 - d) * (
                    1 - e) + a * (1 - b) * c * (1 - d) * (1 - e) + \
                      (1 - a) * b * (1 - c) * d * (1 - e) + (1 - a) * (1 - b) * c * d * (1 - e) + a * (1 - b) * (1 - c) * (
                                  1 - d) * e + (1 - a) * b * (1 - c) * (1 - d) * e + \
                      (1 - a) * (1 - b) * c * (1 - d) * e + (1 - a) * (1 - b) * (1 - c) * d * e)
        round_dic[f'+{c}'] = max_4
        round_dic[f'-{c}'] = min_4
        if c == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['c'] = [max_round, min_round]
    for d in num:
        c = exp ** (-λc * t)
        min_4 = a * b * c * (1 - d) * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (
                1 - a) * b * c * d * (1 - e) \
                  + a * b * c * d * (1 - e) + a * b * (1 - c) * (1 - d) * e + a * (1 - b) * c * (1 - d) * e + (
                              1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e \
                  + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (
                              1 - a) * b * c * d * e + a * b * c * d * e
        max_4 = 1 - (a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e + a * b * (1 - c) * (1 - d) * (
                    1 - e) + a * (1 - b) * c * (1 - d) * (1 - e) + \
                      (1 - a) * b * (1 - c) * d * (1 - e) + (1 - a) * (1 - b) * c * d * (1 - e) + a * (1 - b) * (1 - c) * (
                                  1 - d) * e + (1 - a) * b * (1 - c) * (1 - d) * e + \
                      (1 - a) * (1 - b) * c * (1 - d) * e + (1 - a) * (1 - b) * (1 - c) * d * e)
        round_dic[f'+{d}'] = max_4
        round_dic[f'-{d}'] = min_4
        if d == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['d'] = [max_round, min_round]
    for e in num:
        d = exp ** (-λd * t)
        min_4 = a * b * c * (1 - d) * (1 - e) + a * b * (1 - c) * d * (1 - e) + a * (1 - b) * c * d * (1 - e) + (
                1 - a) * b * c * d * (1 - e) \
                  + a * b * c * d * (1 - e) + a * b * (1 - c) * (1 - d) * e + a * (1 - b) * c * (1 - d) * e + (
                              1 - a) * b * c * (1 - d) * e + a * b * c * (1 - d) * e + a * (1 - b) * (1 - c) * d * e \
                  + (1 - a) * b * (1 - c) * d * e + a * b * (1 - c) * d * e + a * (1 - b) * c * d * e + (
                              1 - a) * b * c * d * e + a * b * c * d * e
        max_4 = 1 -( a * b * (1 - c) * (1 - d) * e + (1 - a) * (1 - b) * c * d * e + a * b * (1 - c) * (1 - d) * (
                    1 - e) + a * (1 - b) * c * (1 - d) * (1 - e) + \
                      (1 - a) * b * (1 - c) * d * (1 - e) + (1 - a) * (1 - b) * c * d * (1 - e) + a * (1 - b) * (1 - c) * (
                                  1 - d) * e + (1 - a) * b * (1 - c) * (1 - d) * e + \
                      (1 - a) * (1 - b) * c * (1 - d) * e + (1 - a) * (1 - b) * (1 - c) * d * e)
        round_dic[f'+{e}'] = max_4
        round_dic[f'-{e}'] = min_4
        if e == 0:
            max_round = round_dic['+1'] - round_dic['-0']
            min_round = round_dic['-1'] - round_dic['+0']
            result_dic['e'] = [max_round, min_round]

    """
    ↑の操作で、偏微分の値を取得。
    その値と、dR(t)/dtの値を
    掛け合わせたものの和を取る。
    故障率の上限と下限を算出
    """
    diff_Rt_Max = 0
    diff_Rt_min = 0
    for idx, link in enumerate(link_list):
        diff_Rt_Max = diff_Rt_Max + result_dic[link][0] * R_diff[idx]
        diff_Rt_min = diff_Rt_min + result_dic[link][1] * R_diff[idx]
    #(Max/minで故障率上限導出)
    approximation_min = (-1/first_min_4)*diff_Rt_Max
    #(min/Maxで故障率上限導出)
    approximation_Max = (-1/first_max_4)*diff_Rt_min
    return approximation_Max, approximation_min, first_max_4, first_min_4

x = sp.Symbol('x')
λt_actual_li = [f_actual(x).subs(x, i) for i in range(20)]
λt_actual = f_actual(x)
t = np.linspace(0, 20)
λt_3_Max, λt_3_min, Rt_3_Max, Rt_3_min = f_3(t)[0], f_3(t)[1], f_3(t)[2], f_3(t)[3]
λt_4_Max, λt_4_min, Rt_4_Max, Rt_4_min = f_4(t)[0], f_4(t)[1], f_4(t)[2], f_4(t)[3]
y1, y2, y3, y4, y5 = λt_actual, λt_3_Max, λt_3_min, λt_4_Max, λt_4_min

wb = xl.Workbook()
ws = wb.active

λt_3_Max = [f_3(t)[0] for t in range(20)]
λt_3_min = [f_3(t)[1] for t in range(20)]
Rt_3_Max = [f_3(t)[2] for t in range(20)]
Rt_3_min = [f_3(t)[3] for t in range(20)]
λt_4_Max = [f_4(t)[0] for t in range(20)]
λt_4_min = [f_4(t)[1] for t in range(20)]
Rt_4_Max = [f_4(t)[2] for t in range(20)]
Rt_4_min = [f_4(t)[3] for t in range(20)]
print(Rt_3_min, Rt_4_min, Rt_4_Max, Rt_3_min)

# ヘッダー作成
header = ['t', 'λt_actual', 'λt_3_Max', 'λt_3_min', 'λt_4_Max', 'λt_4_min', 'Rt_3_max', 'Rt_3_min', 'Rt_4_max', 'Rt_4_min']
for i in range(len(header)):
    ws.cell(1, i+1, value=header[i])
# 配列ループ
for i in range(0, len(λt_3_Max)):
    ws.cell(i + 2, 2, value=f'{λt_actual_li[i]}')
    ws.cell(i + 2, 3, value=λt_3_Max[i])
    ws.cell(i + 2, 4, value=λt_3_min[i])
    ws.cell(i + 2, 5, value=λt_4_Max[i])
    ws.cell(i + 2, 6, value=λt_4_min[i])
    ws.cell(i + 2, 7, value=Rt_3_Max[i])
    ws.cell(i + 2, 8, value=Rt_3_min[i])
    ws.cell(i + 2, 9, value=Rt_4_Max[i])
    ws.cell(i + 2, 10, value=Rt_4_min[i])


dt_now = datetime.datetime.now().strftime('%Y.%m.%d.%H.%M')
filename = f'lab_{dt_now}'
wb.save(f'/Users/ranwatanabe/Desktop/{filename}.xlsx')

fig, ax = plt.subplots()
plt.title('λ(t)Max and λ(t)min')
plt.xlabel('t')
plt.ylabel('Failure rate λ(t)')
ax.plot(t, y2, label='approximation_3_Max')
ax.plot(t, y3, label='approximation_3_min')
ax.plot(t, y4, label='approximation_4_Max')
ax.plot(t, y5, label='approximation_4_min')
ax.plot(t, [sp.lambdify(x, y1)(i) for i in t], label='actual')
ax.legend()
plt.show()


