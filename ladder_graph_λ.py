import networkx as nx
import matplotlib.pyplot as plt
import numpy as np
import sympy as sp
import string
import itertools
from function import stop_watch
from colorama import init, Fore
import datetime
import openpyxl as xl
import copy
init()

# TODO 不信頼度のグラフを描画し、故障率のグラフも描画する。
# TODO 不信頼度の値をエクセルファイルに出力
# TODO 不信頼度の近似値をエクセルファイルに出力
# TODO 現状→故障率のグラフ描画、エクセル出力。
"""
近似時間増加理由①：正常な状態と、故障した状態を全て列挙（近似で使うため。）
近似時間増加理由②：上限加減ともに、代入作業、偏微分近似作業があるため。(感覚的には倍以上→故障している場合が莫大に存在するため)
近似時間増加理由③：変数代入を行っている,failure_rate計算において、折角無視したのに、全ての計算を行っていることと同値であるため。
"""


class GraphLogic:
    def __init__(self, graph_size, t):
        # サイズに合わせて、アルファベットをノードに見立てる。
        self.nord_list = list(string.ascii_lowercase)[0:int(graph_size[0] * graph_size[1])]
        # 変数定義
        self.var = {i: sp.Symbol(f'{i}') for i in self.nord_list}
        # ノードの値設定
        self.λ = {i: sp.E**(-0.001 * t) for i in self.nord_list}
        self.nord_value = {i: {"bool": ["0", "1"], "0": 1 - self.var[i], "1": self.var[i]} for i in self.nord_list}
        self.param = [(self.var[i], self.λ[i]) for i in self.nord_list]

    @stop_watch
    def get_shotest_path(self, a, b) -> list[str]:
        """
        axbのラダーグラフを作成し、K-節点間を結ぶ最短経路を算出
        :param a:列
        :param b:行
        :return: K-節点を結ぶ最短経路のリスト
        """
        # 3x3のラダーグラフを生成
        G = nx.generators.lattice.grid_graph(dim=[a, b])
        # ノードを英字で表記するための辞書を作成
        node_labels = {(i, j): self.nord_list[i * 3 + j] for i in range(b) for j in range(a)}
        # ノードの表記を変更
        G = nx.relabel_nodes(G, node_labels)
        # nx.draw(G)
        # plt.show()
        # 開始ノードと終了ノードを定義(K-節点の終点と始点)
        start_node = self.nord_list[0]
        end_node = self.nord_list[-1]
        # 複数の最短経路を求める
        shortest_paths = [path for path in nx.all_shortest_paths(G, start_node, end_node)]
        return shortest_paths

    @stop_watch
    def get_truth_table(self) -> set[tuple]:
        """
        全てのノードの状態を列挙した真理値表を作成
        :return: 真理値表の集合{(0,0),(0,1),(1,0),(1,1)}
        """
        all_nord_status = []
        all_bool = [self.nord_value[i]["bool"] for i in self.nord_value]
        all_nord_status.extend(list(itertools.product(*all_bool)))
        all_nord_status = set(all_nord_status)
        return all_nord_status

    @stop_watch
    def subs(self, value, var_name):
        print(f"{var_name}パラメータ代入処理時間↓")
        value = value.subs(self.param)
        return value

    def fastest_approximate_status(self, shortest_path_list) -> set[tuple]:
        """
        全列挙を行わないため、最速の処理が可能。
        :param all_nord_status: 全ての状態を示した真理値表
        :param shortest_path_list: 取得した最短経路
        :return: 全ての正常な道の集合
        """
        in_shortest_roots = []
        for shotest_path in shortest_path_list:  # shotest_path→['a', 'd', 'g', 'h', 'i']
            total_bool = [["1"] if i in shotest_path else self.nord_value[i]["bool"] for i in self.nord_value]
            in_shortest_roots.extend([i for i in list(itertools.product(*total_bool))])
        # 重複を含まない、正常状態である道の集合
        in_shortest_roots = set(in_shortest_roots)
        return in_shortest_roots

    @stop_watch
    def actual_status(self, all_nord_status, shortest_path_list) -> set[tuple]:
        """
        真理値表法なので、原則に基づいて算出する（冗長な処理）
        最短経路が真理値表に含まれる場合→正常　　と判定、その後集合に含める(重複削除)
        :param all_nord_status: 全ての状態を示した真理値表
        :param shortest_path_list: 取得した最短経路
        :return: 全ての正常な道の集合
        """
        in_shortest_roots = []
        for i in shortest_path_list:
            idx = [self.nord_list.index(f"{j}") for j in i]  # [1,4,7,8]
            for root in all_nord_status:
                if all([root[i] == "1" for i in idx]):
                    in_shortest_roots.append(root)
        in_shortest_roots = set(in_shortest_roots)
        # in_shortest_rootsは、重複を含まない、正常である道のリスト
        return in_shortest_roots

    @stop_watch
    def approximate_status(self, all_nord_status, shortest_path_list) -> set[tuple]:
        """
        正常な状態と、故障した状態を全て列挙（近似で使うため。）
        :param all_nord_status: 全ての状態を示した真理値表
        :param shortest_path_list: 取得した最短経路
        :return: 全ての正常な道の集合
        """
        in_shortest_roots = []
        for shotest_path in shortest_path_list:  # shotest_path→['a', 'd', 'g', 'h', 'i']
            total_bool = [["1"] if i in shotest_path else self.nord_value[i]["bool"] for i in self.nord_value]
            in_shortest_roots.extend([i for i in list(itertools.product(*total_bool))])
        # 重複を含まない、正常状態である道の集合
        in_shortest_roots = set(in_shortest_roots)
        # 重複を含まない、故障状態である道の集合
        out_shortest_roots = {x for x in all_nord_status if x not in in_shortest_roots}
        return in_shortest_roots, out_shortest_roots

    @stop_watch
    def fastest_approximate_reliability(self, in_shortest, m, reliability=1) -> sp:
        """
        下限のみを算出するため全列挙が不必要かつ、mに応じて演算量削減
        ただし、上限近似を出すことができない。精度はよい。
        :param in_shortest: 正常な道
        :param m: m個以上の故障を計算から除外する識別子
        :param reliability: 初期値
        :return: 近似した信頼度
        """
        min_list = []
        # 下限近似算出
        for root in in_shortest:
            if root.count("0") >= m:
                continue
            else:
                for idx, nord in enumerate(root):
                    reliability *= self.nord_value[self.nord_list[idx]][nord]
                min_list.append(reliability)
                reliability = 1
        min = sum(min_list)
        return min

    @stop_watch
    def actual_reliability(self, truth_root, reliability=1) -> sp:
        """
        1,0,1のようにそれぞれのノードの正常(故障)状態に応じて、値を掛け合わせ、
        全ての道の信頼度の和をとる
        :param truth_root: 正常な道
        :param reliability: 初期値を設定しているだけ
        :return:　グラフ全体の信頼度
        """
        value_list = []
        for root in truth_root:
            for idx, nord in enumerate(root):
                reliability *= self.nord_value[self.nord_list[idx]][nord]
            value_list.append(reliability)
            reliability = 1
        rel = sum(value_list)
        return rel

    @stop_watch
    def approximate_reliability(self, in_shortest, out_shortest, m, reliability=1) -> tuple[sp, sp]:
        """
        下限：mを無視した信頼度
        上限：mを無視した故障を1から引いた値
        :param in_shortest: 正常な道
        :param out_shortest: 故障した道
        :param m: m個以上の故障を計算から除外する識別子
        :param reliability: 近似した信頼度
        :return:
        """
        min_list = []
        # 下限近似算出
        for root in in_shortest:
            if root.count("0") >= m:
                continue
            else:
                for idx, nord in enumerate(root):
                    reliability *= self.nord_value[self.nord_list[idx]][nord]
                min_list.append(reliability)
                reliability = 1
        min = sum(min_list)
        # 上限近似算出
        Max_list = []
        for root in out_shortest:
            if root.count("0") >= m:
                continue
            else:
                for idx, nord in enumerate(root):
                    reliability *= self.nord_value[self.nord_list[idx]][nord]
                Max_list.append(reliability)
                reliability = 1
        Max = 1 - sum(Max_list)
        return min, Max

    @stop_watch
    def actual_diff(self, rel) -> tuple[sp, sp]:
        """
        １変数を1に置き換えたものから0に置き換えたものを引く。
        それを全ての変数で行い、総和をとる。
        その際に代入するRtのminとmaxで調整し、diff_minとdiff_maxを作る
        :param rel: 信頼度
        :return: 信頼度の微分の上限、加減
        """
        round_dic = {}
        result_dic = {}
        # それぞれのRi(t)の微分を算出
        var_diff = [sp.diff(self.λ[i], t) for i in self.nord_list]
        # 偏微分の近似計算
        for idx, i in enumerate([self.var[i] for i in self.nord_list]):
            for value in [1, 0]:
                calc_rel = copy.copy(rel).subs(i, value)
                round_dic[f'{self.nord_list[idx]}{value}'] = calc_rel
        for nord in self.nord_list:
            round = round_dic[f'{nord}1'] - round_dic[f'{nord}0']
            result_dic[f'{nord}'] = round
        diff_Rt = 0
        for idx, nord in enumerate(self.nord_list):
            diff_Rt = diff_Rt + result_dic[nord] * var_diff[idx]
        return diff_Rt, rel

    @stop_watch
    def approximate_diff(self, rel) -> tuple[sp, sp]:
        """
        １変数を1に置き換えたものから0に置き換えたものを引く。
        それを全ての変数で行い、総和をとる。
        その際に代入するRtのminとmaxで調整し、diff_minとdiff_maxを作る
        :param rel: 信頼度
        :return: 信頼度の微分の上限、加減
        """
        round_dic = {}
        result_dic = {}
        # それぞれのRi(t)の微分を算出
        var_diff = [sp.diff(self.λ[i], t) for i in self.nord_list]
        # 偏微分の近似計算
        for idx, i in enumerate([self.var[i] for i in self.nord_list]):
            for value in [1, 0]:
                calc_rel = copy.copy(rel)
                Max = calc_rel["max"].subs(i, value)
                min = calc_rel["min"].subs(i, value)
                round_dic[f'+{self.nord_list[idx]}{value}'] = Max
                round_dic[f'-{self.nord_list[idx]}{value}'] = min
        for nord in self.nord_list:
            max_round = round_dic[f'+{nord}1'] - round_dic[f'-{nord}0']
            # 偏微分の部分の下限(fmin(1)-fMax(0))を導出　　この値が常に正とは限らない。
            min_round = round_dic[f'-{nord}1'] - round_dic[f'+{nord}0']
            result_dic[f'{nord}'] = {"max": max_round, "min": min_round}
        diff_Rt_Max = 0
        diff_Rt_min = 0
        for idx, nord in enumerate(self.nord_list):
            diff_Rt_Max = diff_Rt_Max + result_dic[nord]["max"] * var_diff[idx]
            diff_Rt_min = diff_Rt_min + result_dic[nord]["min"] * var_diff[idx]
        return diff_Rt_min, diff_Rt_Max

    @stop_watch
    def fastest_approximate_failure_rate(self, rel):
        """
        信頼度の変数に実際の値を入れ、値を求める。
        :param rel: 信頼度
        :return: 故障率
        """
        approximation = rel.subs(self.param)
        failure_rate = (-1 / approximation) * sp.diff(approximation, t)
        return failure_rate, approximation

    @stop_watch
    def actual_failure_rate(self, rel, diff) -> sp:
        failure_result = (-1 / rel) * diff
        return failure_result

    @stop_watch
    def approximate_failure_rate(self, rel_min, rel_max, diff_min, diff_max):
        """
        minとmaxがそれぞれの変数にあることから、全体のminとmaxを算出する
        :param rel: 信頼度
        :param diff: 信頼度の微分
        :return:
        """
        # (Max/minで故障率上限導出)
        app_failure_Max = (-1 / rel_min) * diff_max
        # (min/Maxで故障率上限導出)
        app_failure_min = (-1 / rel_max) * diff_min
        return app_failure_min, app_failure_Max,

@stop_watch
def get_actual_failure_rate(a, b, t):
    """
    :param a:ラダーの列
    :param b:ラダーの行
    :return: actualの信頼度　変数tを含む
    """
    #　init作成
    ladder = GraphLogic([a, b], t)
    # 正常であるルートを探索
    truth_table = ladder.get_truth_table()
    shortest_path = ladder.get_shotest_path(a, b)
    shortest_root = ladder.actual_status(truth_table, shortest_path)
    # 信頼度計算
    actual_rel = ladder.actual_reliability(shortest_root)
    # 微分近似計算
    rel_diff, actual_rel = ladder.actual_diff(actual_rel)
    actual_rel = ladder.subs(actual_rel, "act_rel")
    rel_diff = ladder.subs(rel_diff, "rel_diff")
    # パラメータを代入したもので故障率計算
    failure_rate = ladder.actual_failure_rate(actual_rel, rel_diff)
    return actual_rel, failure_rate

@stop_watch
def get_approximate_failure_rate(a, b, t, m) -> dict:
    rel, diff, failure = {}, {}, {}
    ladder = GraphLogic([a, b], t)
    truth_table = ladder.get_truth_table()
    shortest_path = ladder.get_shotest_path(a, b)
    in_root, out_root = ladder.approximate_status(truth_table, shortest_path)
    rel["min"], rel["max"] = ladder.approximate_reliability(in_root, out_root, m)
    diff["min"], diff["max"] = ladder.approximate_diff(rel)
    rel["min"] = ladder.subs(rel["min"], "rel_min")
    rel["max"] = ladder.subs(rel["max"], "rel_max")
    # m個以上の同時故障を無視
    diff["min"] = ladder.subs(diff["min"], "diff_min")
    diff["max"] = ladder.subs(diff["max"], "diff_max")
    failure["min"], failure["max"] = ladder.approximate_failure_rate(rel["min"], rel["max"], diff["min"], diff["max"])
    return rel, failure

@stop_watch
def get_fastest_approximate_failure_rate(a, b, t, m):
    ladder = GraphLogic([a, b], t)
    shortest_path = ladder.get_shotest_path(a, b)
    in_root = ladder.fastest_approximate_status(shortest_path)
    reliability = ladder.fastest_approximate_reliability(in_root, m)  # m個以上の同時故障を無視
    failure, reliability= ladder.fastest_approximate_failure_rate(reliability)
    return reliability, failure

t = sp.Symbol('t')
t_value = np.linspace(0, 100)
# メイン関数実　size
size = [4]
fast = {}
act = {}
app = {}
wb_value = xl.Workbook()

# ファイル名作成
dt_now = datetime.datetime.now().strftime('%Y.%m.%d.%H.%M')
filename_time = f'lab_{dt_now}'
header_time = ['size', '故障無視数(m以上)', '関数名', '処理時間【秒】', '総処理時間【秒】']
for num in size:
    # actualの計算
    print(Fore.GREEN + f'【真値】故障率計算3*{num}')
    act[f"act_rel_3*{num}"], act[f"act_fail_3*{num}"] = get_actual_failure_rate(3, num, t)
    print('-----------------------------------------------------------')
    for j in range(2, 8):
        # act エクセル出力
        ws_value = wb_value.active
        new_sheet = wb_value.create_sheet(f'm={j}')
        ws_value = wb_value[f'm={j}']
        # ヘッダー設定
        header_value = ["t", f'app_m{j}_min', f'3*{num}act', f'app_m{j}_max']
        for i in range(len(header_value)):
            ws_value.cell(1, i + 1, value=header_value[i])
        # 値をエクセルに書き込み
        for i in range(50):
            ws_value.cell(i + 2, 1, str(i))
            ws_value.cell(i + 2, 3, value=str(act[f"act_fail_3*{num}"].subs(t, i)))
        fig_rel, ax_rel = plt.subplots()
        plt.title(f'R(t)Max and R(t)min and actual [3×{num}]')
        plt.xlabel('t')
        plt.ylabel('reliability R(t)')
        # failのグラフ描画設定
        fig_fail, ax_fail = plt.subplots()
        plt.title(f'λ(t)Max and λ(t)min [3×{num}]')
        plt.xlabel('t')
        plt.ylabel('Failure rate λ(t)')
        # actualのプロット
        ax_rel.plot(t_value, [sp.lambdify(t, act[f"act_rel_3*{num}"])(i) for i in t_value], label=f'act_rel_3*{num}')
        ax_fail.plot(t_value, [sp.lambdify(t, act[f"act_fail_3*{num}"])(i) for i in t_value], label=f'act_fail_3*{num}')
        # fastestの計算とプロット
        # print(Fore.BLUE + f'【最速】故障率計算3*{num}、{j}個以上の故障ケース無視')
        # fast[f"fast_rel_{num}"], fast[f"fast_fail_{num}"] = get_fastest_approximate_failure_rate(3, num, t, j)
        # print('-----------------------------------------------------------')
        # ax_rel.plot(t_value, [sp.lambdify(t, fast[f"fast_rel_{num}"])(i) for i in t_value], label=f'fast_rel_{num}')
        # ax_fail.plot(t_value, [sp.lambdify(t, fast[f"fast_fail_{num}"])(i) for i in t_value], label=f'fast_fail_{num}')
        # rel_appの計算とプロット
        print(Fore.YELLOW + f'【近似】故障率計算3*{num}、{j}個以上の故障ケース無視')
        app[f"app_rel_3*{num}_m{j}"], app[f"app_fail_3*{num}_m{j}"] = get_approximate_failure_rate(3, num, t, j)
        print('-----------------------------------------------------------')
        # エクセルのヘッダー設定
        ws_value.cell(1, 2, value=f"app_{j}_min")
        ws_value.cell(1, 4, value=f"app_{j}_max")
        for i in range(50):
            # minとMaxをエクセルに書き込み
            ws_value.cell(i + 2, 2, value=str(app[f"app_fail_3*{num}_m{j}"]["min"].subs(t, i)))
            ws_value.cell(i + 2, 4, value=str(app[f"app_fail_3*{num}_m{j}"]["max"].subs(t, i)))
        ax_rel.plot(t_value, [sp.lambdify(t, app[f"app_rel_3*{num}_m{j}"]["min"])(i) for i in t_value], label=f'app_rel_3*{num}_m{j}_min')
        ax_rel.plot(t_value, [sp.lambdify(t, app[f"app_rel_3*{num}_m{j}"]["max"])(i) for i in t_value], label=f'app_rel_3*{num}_m{j}_max')
        # fail_appの計算とプロット
        ax_fail.plot(t_value, [sp.lambdify(t, app[f"app_fail_3*{num}_m{j}"]["min"])(i) for i in t_value], label=f'app_fail_3*{num}_m{j}_min')
        ax_fail.plot(t_value, [sp.lambdify(t, app[f"app_fail_3*{num}_m{j}"]["max"])(i) for i in t_value], label=f'app_fail_3*{num}_m{j}_max')
        # ax_rel.legend()
        ax_fail.legend()

wb_value.remove(wb_value.worksheets[0])
wb_value.save(f'/Users/ranwatanabe/Desktop/{filename_time}.xlsx')
plt.show()

